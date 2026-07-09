import os
import csv
import pytest
from src.reporter import export_to_csv, export_to_excel, cleanup_temp_dir

@pytest.fixture
def dummy_data():
    return [
        {"nim": "12345", "base_score": 90, "penalty": 0, "final_score": 90, "notes": ""},
        {"nim": "67890", "base_score": 80, "penalty": 2, "final_score": 78, "notes": "REVIEW MANUAL"},
        {"nim": "11111", "base_score": 69, "penalty": 0, "final_score": 69, "notes": "Plagiat"},
        {"nim": "22222", "base_score": 0, "penalty": 0, "final_score": 0, "notes": "Gagal kompilasi"}
    ]

def test_export_to_csv(dummy_data, tmp_path):
    csv_file = tmp_path / "test_report.csv"
    success = export_to_csv(dummy_data, str(csv_file))
    
    assert success is True
    assert os.path.exists(csv_file)
    
    with open(csv_file, mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        rows = list(reader)
        
    assert len(rows) == 5 # 1 header + 4 data
    assert rows[0] == ["NIM", "Nilai Dasar", "Penalti", "Nilai Akhir", "Keterangan"]
    assert rows[1] == ["12345", "90", "0", "90", ""]

def test_export_to_excel(dummy_data, tmp_path):
    excel_file = tmp_path / "test_report.xlsx"
    success = export_to_excel(dummy_data, str(excel_file))
    
    # openpyxl might not be installed, so we check if it succeeds based on availability
    try:
        import openpyxl
        assert success is True
        assert os.path.exists(excel_file)
        
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "NIM"
        assert str(ws.cell(row=2, column=1).value) == "12345"
    except ImportError:
        assert success is False

def test_cleanup_temp_dir(tmp_path):
    temp_dir = tmp_path / "test_temp"
    temp_dir.mkdir()
    
    dummy_file = temp_dir / "dummy.txt"
    dummy_file.write_text("dummy content")
    
    assert os.path.exists(temp_dir)
    assert os.path.exists(dummy_file)
    
    cleanup_temp_dir(str(temp_dir))
    
    assert not os.path.exists(temp_dir)
